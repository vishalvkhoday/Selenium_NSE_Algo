'''
Created on Jul 13, 2018

@author: khoday
'''

import  time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.common.exceptions import NoSuchElementException
import pyautogui
from openpyxl import load_workbook
import os
import datetime


def objExist(obj):
    try:
        obj.is_displayed()
        return True
    except NoSuchElementException:
        return False

def getAbsPath():
    path = os.getcwd()
    return str(path).replace('\\', '/')

def Screenshot():
        timstr = str(time.strftime('%Y%m%d%H%M%S'))
        pic =pyautogui.screenshot()
        scr_path = getAbsPath()
        pic.save(scr_path+'/Screen_shot/img'+timstr+'.png')

def Table_Value (str_val):
    temp_tble_Val = str(str_val).split("\n")
    temp_tble_Val = str(temp_tble_Val[0]).replace("\n", "").split("\n")
    temp_str = ""
    for i in temp_tble_Val:
        if len(str(i).strip()) > 0:
            i = str(i).strip()
            temp_str += i + "\n"
    return temp_str

def clear_Temp ():
    try:
        os.system('del /f/s/q %temp%\*')
    except AssertionError as ae:
        print('Unable to delete temp files')
    
def get_Sector():
    try:
        Head_Inner_Val = driver.find_element_by_class_name('bsns_pcst').get_attribute('innerText')
        splt_header = str(Head_Inner_Val).split('|')
        t_bse_code = str(splt_header[0]).split(':')
        bse_code = str(t_bse_code[1]).strip()
        t_nse_code = str(splt_header[1]).split(':')
        nse_code = str(t_nse_code[1]).strip()    
        t_ISIN = str(splt_header[2]).split(':')
        ISIN = str(t_ISIN[1]).strip()    
        t_sector = str(splt_header[3]).split(":")
        Sector = str(t_sector[1]).strip()
        print (splt_header)    
        return Sector,ISIN,nse_code
    except:
        return 1

def filter_StockInfo(Str_value):    
    if (Str_value.strip()):
        return True
    else:
        False
        
def Convert(lstfull):    
    dict_list = {}
    lst = lstfull.split(",")
    try:
        for i in range(0,len(lst),2):
            print (lst[i],lst[i+1])
            dict_list.update({lst[i]:lst[i+1]})
    
#     res_dct = {lst[i]: lst[i + 1] for i in range(0, len(lst), 2)} 
    except:
        return dict_list    

def Stock_info (str_tbl,f_Name):    
    Ws = Wb['Sheet1']    
    try:
        driver.find_element_by_id(str_tbl).location_once_scrolled_into_view
        t_Tbl_InnerText_Val = str(driver.find_element_by_id(str_tbl).get_attribute('innerText')).strip()
        Tbl_InnerText_Val = t_Tbl_InnerText_Val.splitlines()
        lst_stockInf = filter(filter_StockInfo,Tbl_InnerText_Val)
        lst_varSt_Dts = ""
        for varSt_Dts in lst_stockInf:
            if (varSt_Dts.strip() == 'Deliverables (%)'):
                break
             
            lst_varSt_Dts = lst_varSt_Dts+varSt_Dts.replace(",","").strip()+","
            print(lst_varSt_Dts)
        
        dict_stock_dt = Convert(lst_varSt_Dts)
    except:
        print('Script not found Error!!!')
#         Screenshot()

#     M_cap = str(Tbl_InnerText_Val[1]).replace("MARKET CAP (RS CR)", "").strip()
    M_cap =dict_stock_dt.get('Market Cap (Rs Cr.)') 
    PE = dict_stock_dt.get('P/E')
    Bookvalue =dict_stock_dt.get('Book Value (Rs)')
    Div = dict_stock_dt.get('Dividend (%)')
    Ind_PE = dict_stock_dt.get('Industry P/E')
    EPS = dict_stock_dt.get('EPS (TTM)')
    PC = dict_stock_dt.get('P/C')
    Price_Book = dict_stock_dt.get('Price/Book')
    DivYield = dict_stock_dt.get('Dividend Yield.(%)')
    Face_val = dict_stock_dt.get('Face Value (RS)')
    
    Sector = get_Sector()
    Ws['B'+ str(row)] =   Sector[1]   
    Ws['C' + str(row)] = M_cap
    Ws['D' + str(row)] = EPS
    Ws['E' + str(row)] = PE
    Ws['F' + str(row)] = PC
    Ws['G' + str(row)] = Bookvalue
    Ws['H' + str(row)] = Price_Book
    Ws['I' + str(row)] = Div
    Ws['J' + str(row)] = DivYield
    Ws['K' + str(row)] = Face_val
    Ws['L' + str(row)] = Ind_PE
    Ws['M' + str(row)] = Sector[0]
    Ws['N' + str(row)] = 'No'
#     Wb.save(f_Name)
    Sector=""
    M_cap = ""
    EPS = ""
    PE = ""
    Bookvalue = ""
    Div = ""
    Face_val = ""
    Ind_PE = ""

def ShareHolding(Script_code,f_Name):
    driver.find_element_by_xpath(id_shr_prt).location_once_scrolled_into_view
#     driver.find_element_by_id(id_shr_prt).click()
    driver.set_page_load_timeout(1)
    sh_Prnt_tbl =driver.find_element_by_xpath(id_shr_prt).get_attribute('innerText')
    sh_Prnt_tbl_ary = str(sh_Prnt_tbl).splitlines()
    r_count = ws_Shr.max_row
    r_count = r_count+1
    for i in range(0,len(sh_Prnt_tbl_ary)-1):
        Col = str(sh_Prnt_tbl_ary[i]).split('\t')
        shColLen = len(Col)
        
        if len(Col) ==4 :
            try:
                                    
                ws_Shr['A'+str(r_count)] = str(Script_code)
                ws_Shr['B'+str(r_count)] = Col[0]
                ws_Shr['C'+str(r_count)] = Col[1]
                ws_Shr['D'+str(r_count)] = Col[2]
                ws_Shr['E'+str(r_count)] = Col[3]
                
                Col[0]=""
                Col[1]=""
                Col[2]=""
                Col[3]=""
                
                r_count = r_count+1
#             Script_code=""
            except:
                r_count = r_count+1
                continue
        elif len(Col)==3:
            try:
                
                ws_Shr['A'+str(r_count)] = str(Script_code)
                ws_Shr['B'+str(r_count)] = Col[0]
                ws_Shr['C'+str(r_count)] = Col[1]
                ws_Shr['D'+str(r_count)] = Col[2]
                r_count = r_count+1
#                 Script_code=""
                Col[0]=""
                Col[1]=""
                Col[2]=""
            except:
                r_count = r_count+1
                continue
            
        else:
            try:
                ws_Shr['A'+str(r_count)] = str(Script_code)
                ws_Shr['B'+str(r_count)] = Col[0]
                ws_Shr['C'+str(r_count)] = Col[1]
                ws_Shr['D'+str(r_count)] = Col[2]
                ws_Shr['E'+str(r_count)] = Col[3]
                ws_Shr['F'+str(r_count)] = Col[4]
                    
#                 Script_code=""
                Col[0]=""
                Col[1]=""
                Col[2]=""
                Col[3]=""
                Col[4]=""
                r_count=r_count+1
            except:
                r_count = r_count+1
                continue 
#     Wb.save(f_Name)

def MF_Holding (Script_code,f_Name):
    driver.set_page_load_timeout(5)
    try:
        driver.find_element_by_xpath(x_MF_holding).location_once_scrolled_into_view
        MF_holding_tbl =driver.find_element_by_xpath(x_MF_holding).get_attribute('innerText')
        if (MF_holding_tbl.strip() == "No Mutual Funds Holding the share"):
            return None
        MF_holding_tbl =MF_holding_tbl.replace("Scheme", "").replace("No. Of Shares", "")
        MF_holding_row = MF_holding_tbl.split("\n")
        fil_str =filter(lambda x:len(x.strip())>0,MF_holding_row)
        temp_list =[]
        for x in fil_str:
            x=x.strip()
            print(x)
            temp_list.append(x)
        mf_row = Ws_MF_Holding.max_row
        mf_row = mf_row+1
        for mf_R in temp_list:
            MF_holding_Split = str(mf_R).split("\t")
            Ws_MF_Holding['A'+str(mf_row)] = str(Script_code)
            Ws_MF_Holding['B'+str(mf_row)] = str(MF_holding_Split[0]).strip()
            Ws_MF_Holding['C'+str(mf_row)] = str(MF_holding_Split[1]).strip()
            mf_row =mf_row +1
    #         Script_code =""
            MF_holding_Split[0]=""
            MF_holding_Split[1]=""
    except:
        print("\nNo Mutal funds holding")
    
    
def Financial(Script_code,f_Name):
    Ws_Fin = Wb["Financial"]
    Fin_row_cnt = Ws_Fin.max_row
    Fin_row_cnt =Fin_row_cnt+1
    driver.find_element_by_id(id_fin_prt).click()
    Fin_tbl = str(driver.find_element_by_xpath(x_fin_tbl).get_attribute('innerText').strip())
    fin_tbl_QoQ = Fin_tbl.splitlines()
    for z in range(0,len(fin_tbl_QoQ)):
        ary_tbl_qoq = fin_tbl_QoQ[z].split('\t')
        
        if (len(ary_tbl_qoq)==1):
            Ws_Fin['F'+str(Fin_row_cnt-1)] =str(ary_tbl_qoq[0])
            continue
        
        else:
            try:
                Ws_Fin['A'+str(Fin_row_cnt)] = str(Script_code).strip()
                Ws_Fin['B'+str(Fin_row_cnt)] = str(ary_tbl_qoq[0])
                Ws_Fin['C'+str(Fin_row_cnt)] = str(ary_tbl_qoq[1])
                Ws_Fin['D'+str(Fin_row_cnt)] =str(ary_tbl_qoq[2])
                Ws_Fin['E'+str(Fin_row_cnt)] =str(ary_tbl_qoq[3])
                Ws_Fin['F'+str(Fin_row_cnt)] =str(ary_tbl_qoq[4])
                
            except:
                continue
        Fin_row_cnt = Fin_row_cnt+1

def Bal_Sheet(Script_code,f_Name):
    Bal_rown_cnt = Ws_Bal_sheet.max_row
    Bal_rown_cnt =Bal_rown_cnt+1
    driver.find_element_by_xpath(x_Bal_Sheet).location_once_scrolled_into_view
    Bal_Tbl = str(driver.find_element_by_xpath(x_Bal_Sheet).get_attribute('innerText')).strip()
    ary_Bal_Tbl = Bal_Tbl.splitlines()
    for y in ary_Bal_Tbl:
        ary_y= str(y).split('\t')
        Ws_Bal_sheet['A'+str(Bal_rown_cnt)]= str(Script_code)
        Ws_Bal_sheet['B'+str(Bal_rown_cnt)]= str(ary_y[0])
        Ws_Bal_sheet['C'+str(Bal_rown_cnt)]= str(ary_y[1])
        Ws_Bal_sheet['D'+str(Bal_rown_cnt)]= str(ary_y[2])
        
        ary_y=""
        Bal_rown_cnt=Bal_rown_cnt+1
    Script_code=""
#     Wb.save(f_Name)
    
getAbsPath()
x_M_cap = '//*[@id="mktdet_1"]/div[1]/div[1]/div[2]'
x_EPS = '//*[@id="mktdet_1"]/div[2]/div[1]/div[2]'
x_PE = '//*[@id="mktdet_1"]/div[1]/div[2]/div[2]'
x_BookValue = '//*[@id="mktdet_1"]/div[1]/div[3]/div[2]'
x_PriceBook = '//*[@id="mktdet_1"]/div[2]/div[3]/div[2]'
x_Div = '//*[@id="mktdet_1"]/div[1]/div[4]/div[2]'
x_Div_yeild = '//*[@id="mktdet_1"]/div[2]/div[4]/div[2]'
x_Facevalue = '//*[@id="mktdet_1"]/div[2]/div[5]/div[2]'
x_Ind_Pe = '//*[@id="mktdet_1"]/div[1]/div[6]/div[2]'
x_error_msg = '//*[@id="mc_mainWrapper"]/div[3]/div[2]/div/p[1]'
x_src_error_msg = '//*[@id="mc_mainWrapper"]/div[3]/div[2]/div/div[3]/p/strong'
x_promo_link = '//*[@id="newsn"]/div/div[2]/p/a'
x_error_msg_tag = '//*[@id="mc_mainWrapper"]/div[3]/div[2]/div/div[3]/p/strong'
f_Name = 'C:/Users/DELL/git/Selenium_NSE_Algo/Additonal_Utility/NSE_Script_codes26Nov2019.xlsx'
x_shr_tbl = '//*[@id="acc_hd7"]/div/div[1]/table'
id_shr_prt = '//table[@class="mctable1 thborder sharePriceTotalCal"]'
id_fin_prt = 'acc_pm5'

x_MF_holding = '(//div/div/table[@class="mctable1 thborder"])[2]'
x_fin_tbl = '//*[@id="findet_1"]/table'
x_Bal_Sheet = '//*[@class="mctable1 thborder"]'
x_bal_dur = '//*[@id="findet_11"]/div/div[2]/div'

Wb = load_workbook(f_Name)
Wb.get_sheet_names()
Ws = Wb['Sheet1']
sht1_Row = Ws.max_row
print (sht1_Row)
# ws_Shr =Wb.get_sheet_by_name("Share_pattern")
ws_Shr =Wb["Share_pattern"]
# Wsheet_partner= Wb.active
r_count = ws_Shr.max_row

if r_count ==1 :
    r_count=2

# Ws_MF_Holding =Wb.get_sheet_by_name("MF_Holding")
Ws_MF_Holding =Wb["MF_Holding"]
mf_row =Ws_MF_Holding.max_row 
if mf_row ==1 :
    mf_row=2
    
Ws_Bal_sheet = Wb["Bal_Sheet"]
row = 1
int_cnt = 1

options = webdriver.ChromeOptions()
options.add_argument("disable-infobars")
options.add_argument("start-maximized")
# driver = webdriver.Chrome(chrome_options=chrome_options)
driver = webdriver.Chrome(chrome_options=options,executable_path='C:/Users/DELL/git/Selenium_NSE_Algo/Additonal_Utility/chromedriver_242', service_args=["--verbose", "--log-path=C:/Users/DELL/git/Selenium_NSE_Algo/Additonal_Utility/Script.log","w+"])

try:
    driver.get('https://www.moneycontrol.com/')
#     driver.get('https://www.moneycontrol.com/india/stockpricequote/mining-minerals/20microns/2M')
    driver.find_element_by_xpath(x_promo_link).click()
except:
    pass
    
driver.set_page_load_timeout(5)
driver.maximize_window()

for row in range(2,sht1_Row):
    Col_Script_code = 'A' + str(row)
    Col_INIE = 'B' + str(row)
    Col_status = 'N' + str(row)
    Script_code = Ws[Col_Script_code].value
    INIE = Ws[Col_INIE].value
    Exe_status = Ws[Col_status].value
    print(Script_code)

    if str(Exe_status).upper() == 'YES': 
        print ('Row number : '+ str(row))         
#         driver.find_element_by_id("search_str").send_keys(INIE)
        obj_Scr_txt =driver.find_element_by_xpath("//*[@id='search_str']")
        obj_Scr_txt.clear()
        obj_Scr_txt.send_keys(Script_code)
#         driver.find_element_by_id("search_str").clear()
#         driver.find_element_by_id("search_str").send_keys(Script_code)
        time.sleep(1)
#         driver.find_element_by_id("search_str").send_keys(Keys.RETURN)
#         os.system('C:/Users/DELL/git/Selenium_NSE_Algo/Additonal_Utility/Enter.vbs')
        ent_path = getAbsPath().replace('/Scripts', '')
#         print(ent_path+'/Additonal_Utility/Enter.vbs')
        os.system(ent_path+'/Additonal_Utility/Enter.vbs')
        ActionChains(driver).send_keys(Keys.ENTER)        
        time.sleep(6)
        driver.set_page_load_timeout(20)

        try:
            str_no_Com = ""
            str_no_Com = driver.find_element_by_xpath(x_error_msg_tag).is_displayed()
            if (str_no_Com == True):
#             str_no_Com = driver.find_element_by_xpath(x_error_msg_tag).get_attribute('innerText')
                driver.find_element_by_id("search_str").send_keys(INIE)
                time.sleep(1)
                ActionChains(driver).send_keys(Keys.ENTER)
                os.system(ent_path+'/Additonal_Utility/Enter.vbs')
        except:
            print('No error while loading page')
        if int_cnt >= 25:
            clear_Temp ()
            int_cnt = 0
        else:
            int_cnt = int_cnt + 1

        try:
            str_sector = get_Sector()
            if str(str_sector[1]).strip() != str(INIE).strip():
                continue
        except:
            continue
        
        try:
            # driver.find_element_by_id('mktdet_nav_2').get_attribute('innerText')
            objValuation =driver.find_element_by_xpath('(//*[contains(text(),"Standalone")])[1]')            
            if objExist(objValuation) ==True:
                Stock_info('standalone_valuation',f_Name)
                time.sleep(2)
            else:
                Stock_info('consolidated_valuation',f_Name)
                time.sleep(2)
                driver.set_page_load_timeout(2)
        except:
            continue
        str_error_msg = ""           
        time.sleep(2)
        driver.set_page_load_timeout(1)
        try:                
#             Financial(Script_code,f_Name)                 
            Bal_Sheet(Script_code,f_Name)                
            ShareHolding(Script_code,f_Name)
            MF_Holding(Script_code,f_Name)
            print (datetime.datetime.now())
            Ws['N' + str(row)] = 'No'
        except:
#             Screenshot()
            print('unknown error occured')
            Wb.save(f_Name)
            continue

        Wb.save(f_Name)
        