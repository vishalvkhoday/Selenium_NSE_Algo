from openpyxl import *
import openpyxl
import pandas


# #strFile = 'C:/Users/Vishal/OneDrive/Desktop/Bse_Results.xls'


# class MergeExcel():

#     def __init__(self,strFile):
#         self.strFile = strFile
        

#     def ObjExcelFile(self):
#         wb=openpyxl.load_workbook(self.strFile,data_only=False,keep_vba=True)
#         objSheet =wb.active
#         maxCol = objSheet.max_column
        
#         maxRow = objSheet.max_row
#         print(maxCol,maxRow)
#         for i in range(1,maxCol):
#             for x in range(1,maxRow):
#                 strRow =objSheet.cell(i,x)

#                 print(strRow.value)



# inputFile = MergeExcel('C:/Users/Vishal/OneDrive/Desktop/PROFORMA_INVOICE_769823.xlsx')
# inputFile.ObjExcelFile()

