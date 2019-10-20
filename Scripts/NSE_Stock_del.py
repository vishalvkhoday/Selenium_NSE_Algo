'''
Created on Aug 11, 2018

@author: khoday
'''
from selenium import webdriver
# import unittest
from datetime import date
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import NoAlertPresentException as NoAlert
# from Algo_Patterns import DB_operation
from selenium.webdriver.support import expected_conditions as EC
import os
import time
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
c_path = os.getcwd()


xl_F_name ="C:/Users/DELL/git/Selenium_NSE_Algo/Additonal_Utility/Down_Scripts_List.xlsx"
Wb =load_workbook(xl_F_name)
Ws=Wb.get_sheet_by_name('Sheet1')

scr_row = Ws.max_row
# db = DB_operation('EXEC     [dbo].[Alog_List_Stock]')
# script =db.connect_DB()
options = webdriver.ChromeOptions()
options.add_argument("start-maximized")
options.add_argument("disable-infobars")

driver = webdriver.Chrome( chrome_options=options,executable_path='C:/Users/DELL/git/Selenium_NSE_Algo/Additonal_Utility/chromedriver_235')
nse_Del_url = 'https://www.nseindia.com/products/content/equities/equities/eq_security.htm'
x_txt_symbol = '//*[@id="symbol"]'
x_bt_GetData = '//*[@id="get"]'
x_rd_bt_duration ='//*[@id="rdDateToDate"]'
x_cal_From_Dt = '//*[@id="fromDate"]'
x_cal_To_Dt = '//*[@id="toDate"]'
x_cal_From_Dt_month = '//*[@id="ui-datepicker-div"]/div/div/select[1]'
x_cal_text = '//*[@id="fromDate"]'
x_lk_dt_file = '//*[@id="historicalData"]/div[1]/span[2]/a'

end_Dt= str(date.today()).split('-')
driver.maximize_window()
driver.get(nse_Del_url)
driver.set_page_load_timeout(5)
driver.find_element_by_xpath(x_rd_bt_duration).click()
pic_dt_from =driver.find_element_by_xpath(x_cal_From_Dt)
# pic_dt_from.send_keys('08-09-2018')
pic_dt_from.send_keys('03-08-2019')
pic_dt_to = driver.find_element_by_xpath(x_cal_To_Dt)
pic_dt_to.send_keys(end_Dt[2]+'-'+end_Dt[1]+'-'+end_Dt[0])

# pic_dt_to.send_keys('03-08-2019')

for x in range(2,scr_row+1):
    try:
        script = Ws[str('A' + str(x))].value
        exe_st = Ws[str('C' +str(x) )].value
        if exe_st =='Yes':            
            driver.find_element_by_xpath(x_txt_symbol).click()
            driver.find_element_by_xpath(x_txt_symbol).clear()
            driver.find_element_by_xpath(x_txt_symbol).send_keys(script)    
            driver.find_element_by_xpath(x_bt_GetData).click()
            driver.set_page_load_timeout(4)
            time.sleep(3)
            driver.find_element_by_xpath(x_lk_dt_file).click()
            Ws['C' +str(x) ]= str('No')
            print ('{}) {} download completed remaining {}'.format(x-1,script,scr_row -x))            
            Wb.save(xl_F_name)
        else:
            print("{})  {} script already downloaded !!!".format(x-1,script))
            continue
    except:
        try:
            if (EC.alert_is_present()):
                Alert(driver).accept()
        except NoAlert:
            pass
        else:
            print("{}) - remaining {} could not be downloaded re run!!!".format(x-1,scr_row -x))
            continue
    
time.sleep(10)
print('Done')


