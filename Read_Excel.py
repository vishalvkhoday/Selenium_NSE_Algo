
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

f_Name = 'c:\\sample.xlsx'


wb = load_workbook(f_Name)
# Sh_name =wb.get_sheet_names()

# grab the active worksheet
# wb.create_sheet('Share_Holding',0)
ws = wb.get_sheet_by_name('Share_Holding')
r_count = ws.max_row
r_count = r_count+1
ws = wb.active
r_nos = str('b'+ str(r_count))
ws[r_nos] = '42'
print wb.get_active_sheet()
print wb.get_sheet_names()
# Data can be assigned directly to cells
# ws['A1'] = 42

# Rows can also be appended
# ws.append([1, 2, 3])

# Python types will automatically be converted
# import datetime
# ws['A3'] = datetime.datetime.now()
# wb.create_sheet("Test", 1)


# Save the file
wb.save(f_Name)

print "done"

